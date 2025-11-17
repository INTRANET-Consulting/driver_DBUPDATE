import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime, date, timedelta
from typing import Dict, List, Tuple, Optional, Any
import re


class ExcelParser:
    """
    Parses weekly planning Excel file with 4 sheets:
    1. Dienste (Routes)
    2. Lenker (Drivers)
    3. Feiertag (Public Holidays)
    4. Dienstplan (Weekly Planning Grid) - DP-Vorlage
    """
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.data = {
            'routes': [],
            'drivers': [],
            'public_holidays': [],
            'driver_availability': [],
            'fixed_assignments': [],
            'school_days': {}
        }
        self.seasonal_routes = {}
        
        print(f"📊 Available sheets in Excel: {self.workbook.sheetnames}")
    
    def _find_sheet(self, *possible_names):
        """Find sheet by multiple possible names (case-insensitive)"""
        sheet_names_lower = {name.lower(): name for name in self.workbook.sheetnames}
        
        for name in possible_names:
            name_lower = name.lower()
            if name_lower in sheet_names_lower:
                return self.workbook[sheet_names_lower[name_lower]]
        
        return None
    
    def parse_all(self, week_start: date) -> Dict[str, Any]:
        """Parse all sheets and return structured data - FIXED ORDER"""
        
        # 1. Parse Lenker sheet FIRST to get driver base info
        print("📋 Step 1: Parsing Lenker sheet...")
        self.parse_lenker_sheet()
        
        # 2. Parse Feiertag sheet BEFORE routes
        print("📋 Step 2: Parsing Feiertag sheet...")
        self.parse_feiertag_sheet()
        
        # 3. Parse Dienstplan sheet to get school days (will also parse driver hours now)
        print("📋 Step 3: Parsing Dienstplan sheet...")
        self.parse_dienstplan_sheet(week_start)
        
        # 4. Parse Dienste sheet - now with day-by-day logic
        print("📋 Step 4: Parsing Dienste sheet...")
        self.parse_dienste_sheet(week_start)
        
        # 5. Parse fixed assignments from Lenker sheet
        print("📋 Step 5: Parsing fixed assignments...")
        self.parse_fixed_assignments(week_start)
        
        return self.data
    
    # ============= HELPER METHOD FOR FUZZY DRIVER MATCHING =============
    
    def _find_matching_driver(self, search_name: str) -> Optional[Dict]:
        """Find matching driver with fuzzy name matching"""
        search_name_lower = search_name.lower().strip()
        
        # Try exact match first
        for driver in self.data['drivers']:
            if driver['name'].lower().strip() == search_name_lower:
                return driver
        
        # Try partial match (handles spelling variations)
        search_parts = search_name_lower.replace(',', '').split()
        
        for driver in self.data['drivers']:
            driver_name_lower = driver['name'].lower().replace(',', '')
            driver_parts = driver_name_lower.split()
            
            # Check if at least 2 parts match (or all parts if name is short)
            matches = sum(1 for part in search_parts 
                         if any(part in dp or dp in part for dp in driver_parts))
            
            required_matches = min(2, len(search_parts))
            
            if matches >= required_matches:
                print(f"   🔄 Fuzzy matched '{search_name}' → '{driver['name']}'")
                return driver
        
        return None
    
    # ============= SHEET 4: DIENSTPLAN (WEEKLY PLANNING) - FIXED VERSION =============
    
    def parse_dienstplan_sheet(self, week_start: date):
        """Parse weekly planning sheet - FIXED to properly detect school days"""
        sheet = self._find_sheet('DP-Vorlage', 'Dienstplan', 'Planning', 'dienstplan', 'Schedule')
        
        if not sheet:
            print("⚠️ Planning sheet not found")
            return
        
        print("📅 Parsing Dienstplan sheet for school days...")
        
        # Strategy: Look for a row with actual dates (not just year)
        date_row = None
        date_start_col = None
        school_status_row = None
        
        # Search more carefully for the calendar section
        for row_idx in range(1, 40):
            for col_idx in range(1, 30):
                cell_value = sheet.cell(row_idx, col_idx).value
                
                if not cell_value:
                    continue
                
                # Check if this is actually a date object (not just a string or number)
                if isinstance(cell_value, (datetime, date)):
                    # Check if the next cell is also a date (confirms we're in date row)
                    next_cell = sheet.cell(row_idx, col_idx + 1).value
                    if isinstance(next_cell, (datetime, date)):
                        date_row = row_idx
                        date_start_col = col_idx
                        # School status is typically 2-3 rows above the date row
                        for offset in [2, 3, 1]:
                            test_row = row_idx - offset
                            test_cell = sheet.cell(test_row, col_idx).value
                            if test_cell and isinstance(test_cell, str):
                                test_str = str(test_cell).lower()
                                if any(keyword in test_str for keyword in ['schul', 'frei', 'ms', 'os']):
                                    school_status_row = test_row
                                    break
                        
                        if not school_status_row:
                            school_status_row = row_idx - 1
                        
                        print(f"✅ Found date row at row {date_row}, starting col {date_start_col}")
                        print(f"   School status row: {school_status_row}")
                        break
                
                # Alternative: Look for "Datum" label
                cell_str = str(cell_value).strip().lower()
                if 'datum' in cell_str and not date_row:
                    test_date_col = col_idx + 1
                    test_cell = sheet.cell(row_idx, test_date_col).value
                    if isinstance(test_cell, (datetime, date)):
                        date_row = row_idx
                        date_start_col = test_date_col
                        school_status_row = row_idx - 1
                        print(f"✅ Found 'Datum' label at row {row_idx}, col {col_idx}")
                        print(f"   Date column starts at: {date_start_col}")
                        print(f"   School status row: {school_status_row}")
                        break
            
            if date_row:
                break
        
        if not date_row or not date_start_col:
            print("⚠️ Could not find date row in Dienstplan")
            print("🔍 Will use AI to determine school vacation periods...")
            self._determine_school_days_with_ai(week_start)
            # Still try to parse driver hours
            self._parse_driver_hours(sheet)
            return
        
        # Parse dates and school status
        dates_parsed = 0
        dates_found = {}
        consecutive_blanks = 0
        
        print(f"📊 Scanning columns starting from {date_start_col}...")
        
        for col_offset in range(0, 100):
            current_col = date_start_col + col_offset
            
            date_cell = sheet.cell(date_row, current_col).value
            
            current_date = None
            if isinstance(date_cell, datetime):
                current_date = date_cell.date()
            elif isinstance(date_cell, date):
                current_date = date_cell
            elif isinstance(date_cell, str):
                current_date = self._parse_date(date_cell)
            
            if not current_date:
                consecutive_blanks += 1
                if dates_parsed > 0 and consecutive_blanks >= 10:
                    break
                continue
            else:
                consecutive_blanks = 0
            
            school_cell = sheet.cell(school_status_row, current_col).value
            
            is_school_day = True
            school_cell_value = ""
            
            if school_cell:
                school_cell_value = str(school_cell).strip()
                school_text = school_cell_value.lower()
                
                if any(keyword in school_text for keyword in [
                    'frei', 'schulfrei', 'ohne', 'ferien', 'vacation', 'holiday', 'os', 'o.s'
                ]):
                    is_school_day = False
                elif any(keyword in school_text for keyword in [
                    'schul', 'ms', 'm.s', 'mit'
                ]):
                    is_school_day = True
            
            dates_found[current_date] = (is_school_day, school_cell_value)
            dates_parsed += 1
        
        print(f"✅ Parsed school days for {dates_parsed} dates")
        
        if dates_found:
            date_range = f"{min(dates_found.keys())} to {max(dates_found.keys())}"
            print(f"📊 Date range found in Excel: {date_range}")
            all_dates = sorted(dates_found.keys())
            sample_dates = all_dates[:3] + ['...'] + all_dates[-3:] if len(all_dates) > 6 else all_dates
            print(f"   Sample dates: {', '.join(str(d) for d in sample_dates)}")
        
        print(f"🎯 Looking for week: {week_start} to {week_start + timedelta(days=6)}")
        for day_offset in range(7):
            current_date = week_start + timedelta(days=day_offset)
            
            if current_date in dates_found:
                is_school_day, cell_value = dates_found[current_date]
                self.data['school_days'][current_date] = is_school_day
                status = "MIT SCHULE" if is_school_day else "OHNE SCHULE"
                print(f"   ✅ {current_date}: {status} ('{cell_value}')")
            else:
                print(f"   ❌ {current_date}: Not found in Excel")
        
        missing_days = []
        for day_offset in range(7):
            current_date = week_start + timedelta(days=day_offset)
            if current_date not in self.data['school_days']:
                missing_days.append(current_date)
        
        if missing_days:
            print(f"⚠️ Missing {len(missing_days)} days from Excel data")
            print("🔍 Using fallback for missing dates...")
            self._determine_school_days_with_ai(week_start)
        
        # Parse driver hours section
        self._parse_driver_hours(sheet)
    
    def _parse_driver_hours(self, sheet: Worksheet):
        """Parse driver worked hours section from Dienstplan"""
        print("\n📊 Parsing driver worked hours from Dienstplan...")
        
        driver_header_row = None
        lenker_col = None
        ist_std_col = None
        
        # Search for the header row with "Lenker" and "Ist-Std"
        for row_idx in range(1, 30):
            for col_idx in range(1, 15):
                cell_value = sheet.cell(row_idx, col_idx).value
                cell_str = str(cell_value).strip() if cell_value else ""
                
                if 'Lenker' in cell_str:
                    driver_header_row = row_idx
                    lenker_col = col_idx
                    print(f"✅ Found 'Lenker' header at row {row_idx}, col {col_idx}")
                    
                    # Now find "Ist-Std" column in the same row
                    for search_col in range(lenker_col, lenker_col + 10):
                        header_cell = sheet.cell(row_idx, search_col).value
                        header_str = str(header_cell).strip() if header_cell else ""
                        
                        if 'ist' in header_str.lower() and 'std' in header_str.lower():
                            ist_std_col = search_col
                            print(f"✅ Found 'Ist-Std' column at col {search_col} ('{header_str}')")
                            break
                    break
            
            if driver_header_row:
                break
        
        if not driver_header_row:
            print("⚠️ Could not find 'Lenker' header in Dienstplan")
            return
        
        if not ist_std_col:
            print("⚠️ Could not find 'Ist-Std' header")
            return
        
        # Parse driver hours
        hours_updated = 0
        drivers_not_found = []
        
        print(f"\n📋 Parsing driver hours starting from row {driver_header_row + 1}...")
        
        for row_idx in range(driver_header_row + 1, min(driver_header_row + 100, sheet.max_row + 1)):
            driver_name = sheet.cell(row_idx, lenker_col).value
            
            if not driver_name or driver_name == '':
                break
            
            driver_name_str = str(driver_name).strip()
            
            # Stop at summary/legend rows
            stop_keywords = [
                'legende', 'dienst', 'summe', 'vollzeit', 
                'feiertag', 'kranken', 'gesamt', 'total'
            ]
            
            # Check if this is a summary row (must match EXACT or be part of compound word)
            if any(driver_name_str.lower() == keyword or 
                   driver_name_str.lower().startswith(keyword + ' ') or
                   driver_name_str.lower().endswith(' ' + keyword)
                   for keyword in stop_keywords):
                print(f"   ⛔ Stopping at row {row_idx}: '{driver_name_str}'")
                break
            
            # Get Ist-Std value
            ist_std_raw = sheet.cell(row_idx, ist_std_col).value
            ist_std = self._parse_time_to_hours(ist_std_raw)
            
            # Skip only if truly empty (not if it's 00:00)
            if ist_std is None:
                print(f"   ⏭️  {driver_name_str}: No hours data")
                continue
            
            # Find matching driver with fuzzy matching
            matching_driver = self._find_matching_driver(driver_name_str)
            
            if matching_driver:
                target = matching_driver['details']['monthly_hours_target']
                remaining = self._subtract_time(target, ist_std) if target and ist_std else None
                
                matching_driver['details']['monthly_hours_worked'] = ist_std
                matching_driver['details']['monthly_hours_remaining'] = remaining
                
                if ist_std == "00:00":
                    print(f"   ✅ {driver_name_str}: worked={ist_std}, target={target}, remaining={remaining} (zero hours)")
                else:
                    print(f"   ✅ {driver_name_str}: worked={ist_std}, target={target}, remaining={remaining}")
                hours_updated += 1
            else:
                drivers_not_found.append(driver_name_str)
                print(f"   ⚠️  {driver_name_str}: Not found in Lenker sheet")
        
        print(f"\n📊 Hours parsing complete:")
        print(f"   ✅ Updated: {hours_updated} drivers")
        if drivers_not_found:
            print(f"   ⚠️  Not found: {', '.join(drivers_not_found[:5])}")
            if len(drivers_not_found) > 5:
                print(f"       ... and {len(drivers_not_found) - 5} more")
    
    def _determine_school_days_with_ai(self, week_start: date):
        """Use typical Austrian school calendar as fallback"""
        print("📚 Using typical Austrian school calendar heuristics...")
        self._use_typical_school_calendar(week_start)
    
    def _use_typical_school_calendar(self, week_start: date):
        """Use typical Austrian school vacation periods as fallback"""
        year = week_start.year
        
        typical_vacations = []
        typical_vacations.append((date(year, 7, 1), date(year, 9, 7)))
        typical_vacations.append((date(year, 12, 24), date(year + 1, 1, 6)))
        typical_vacations.append((date(year - 1, 12, 24), date(year, 1, 6)))
        typical_vacations.append((date(year, 2, 1), date(year, 2, 7)))
        typical_vacations.append((date(year, 10, 26), date(year, 11, 2)))
        
        print(f"📅 Checking week against typical vacation periods...")
        
        for day_offset in range(7):
            current_date = week_start + timedelta(days=day_offset)
            is_vacation = False
            
            for vacation_start, vacation_end in typical_vacations:
                if vacation_start <= current_date <= vacation_end:
                    is_vacation = True
                    break
            
            is_school_day = not is_vacation
            self.data['school_days'][current_date] = is_school_day
            
            status = "MIT SCHULE" if is_school_day else "OHNE SCHULE (vacation - estimated)"
            print(f"   📅 {current_date}: {status}")
    
    # ============= DIENSTE SHEET PARSING =============
    
    def parse_dienste_sheet(self, week_start: date):
        """Parse routes sheet with day-by-day seasonal availability"""
        sheet = self._find_sheet('Dienste', 'Routes', 'dienste')
        
        if not sheet:
            print("⚠️ Routes sheet not found")
            return
        
        print("📋 Parsing Dienste sheet...")
        
        route_definitions = self._parse_route_definitions(sheet)
        print(f"✅ Parsed {len(route_definitions)} route definitions")
        
        self._parse_seasonal_routes(sheet)
        
        print(f"📅 Generating routes for week starting {week_start}...")
        self._generate_weekly_routes(week_start, route_definitions)
    
    def _parse_route_definitions(self, sheet: Worksheet) -> Dict[str, Dict]:
        """Parse first table with route details"""
        route_defs = {}
        start_row = 3
        
        for row_idx in range(start_row, min(start_row + 100, sheet.max_row + 1)):
            row = sheet[row_idx]
            dienst_nr = row[1].value
            
            if not dienst_nr or dienst_nr in ['Dienst-Nr.', 'Dienst-Nr']:
                continue
            
            dienst_nr_str = str(dienst_nr).strip()
            
            if dienst_nr_str in ['FT', 'K', 'FREI', 'F', 'U']:
                continue
            
            if dienst_nr_str in ['DI', 'MB', 'SOF']:
                route_defs[dienst_nr_str] = {
                    'linien_dienst': row[0].value,
                    'dienst_nr': dienst_nr_str,
                    'vad_mit_schule': None,
                    'vad_ohne_schule': None,
                    'diaten': None,
                    'tag': None,
                    'kfz_ort': None,
                    'is_special_duty': True
                }
                continue
            
            route_defs[dienst_nr_str] = {
                'linien_dienst': row[0].value,
                'dienst_nr': dienst_nr_str,
                'vad_mit_schule': self._parse_time(row[2].value),
                'vad_ohne_schule': self._parse_time(row[3].value),
                'diaten': self._parse_number(row[4].value),
                'tag': row[5].value,
                'kfz_ort': row[6].value,
                'is_special_duty': False
            }
        
        return route_defs
    
    def _parse_seasonal_routes(self, sheet: Worksheet) -> None:
        """Parse seasonal route availability table"""
        self.seasonal_routes = {
            'summer_mit_schule': [],
            'summer_ohne_schule': [],
            'winter_mit_schule': [],
            'winter_ohne_schule': []
        }
        
        header_row = None
        start_col = 8
        
        for row_idx in range(1, 10):
            row = sheet[row_idx]
            cell_val = str(row[start_col].value).strip() if row[start_col].value else ''
            if cell_val in ['SmS', 'SoS', 'WmS', 'WoS']:
                header_row = row_idx
                break
        
        if not header_row:
            print("⚠️ Could not find seasonal routes table")
            return
        
        print(f"✅ Found seasonal routes table at row {header_row}")
        
        col_mapping = {
            start_col: 'summer_mit_schule',
            start_col + 1: 'summer_ohne_schule',
            start_col + 2: 'winter_mit_schule',
            start_col + 3: 'winter_ohne_schule'
        }
        
        for col_idx, season_key in col_mapping.items():
            routes_all = []
            
            for row_idx in range(header_row + 1, min(header_row + 100, sheet.max_row + 1)):
                cell_value = sheet.cell(row_idx, col_idx + 1).value
                
                if not cell_value or str(cell_value).strip() == '':
                    continue
                
                route_name = str(cell_value).strip()
                
                if route_name in ['FT', 'K', 'F', 'U', 'SOF']:
                    continue
                
                routes_all.append(route_name)
            
            routes_base_only = self._filter_to_base_routes(routes_all)
            self.seasonal_routes[season_key] = routes_base_only
        
        for season_key, routes in self.seasonal_routes.items():
            print(f"  {season_key}: {len(routes)} routes")
    
    def _filter_to_base_routes(self, routes: List[str]) -> List[str]:
        """Filter to base routes only"""
        base_routes_in_list = set()
        vor_nach_variants = {}
        other_routes = []
        
        for route in routes:
            if route.endswith('-vor'):
                base_route = route[:-4]
                if base_route not in vor_nach_variants:
                    vor_nach_variants[base_route] = []
                vor_nach_variants[base_route].append(route)
            elif route.endswith('-nach'):
                base_route = route[:-5]
                if base_route not in vor_nach_variants:
                    vor_nach_variants[base_route] = []
                vor_nach_variants[base_route].append(route)
            else:
                base_routes_in_list.add(route)
                other_routes.append(route)
        
        filtered = []
        filtered.extend(other_routes)
        
        for base_route, variants in vor_nach_variants.items():
            if base_route not in base_routes_in_list:
                filtered.append(base_route)
        
        return filtered
    
    def _get_seasonal_routes(self, season_key: str) -> List[str]:
        """Get routes for specific season+school combination"""
        return self.seasonal_routes.get(season_key, [])
    
    def _generate_weekly_routes(self, week_start: date, route_definitions: Dict):
        """Generate route entries for each day"""
        holiday_dates = set()
        for holiday in self.data['public_holidays']:
            if week_start <= holiday['date'] < week_start + timedelta(days=7):
                holiday_dates.add(holiday['date'])
        
        for day_offset in range(7):
            current_date = week_start + timedelta(days=day_offset)
            day_name = current_date.strftime('%A')
            
            if current_date in holiday_dates:
                holiday_name = next((h['name'] for h in self.data['public_holidays'] 
                                   if h['date'] == current_date), 'Holiday')
                print(f"  🎉 {current_date} ({day_name}): {holiday_name} - NO ROUTES")
                continue
            
            season = self._get_season_for_date(current_date)
            is_school_day = self.data['school_days'].get(current_date, True)
            school_status = 'mit_schule' if is_school_day else 'ohne_schule'
            
            print(f"  📅 {current_date} ({day_name}): {season}, {school_status}")
            
            if day_offset == 6:
                print(f"    ⭐ Sunday - no routes")
                continue
            
            season_key = self._get_season_key(season, school_status)
            applicable_routes = self._get_seasonal_routes(season_key)
            
            if day_offset == 5:  # Saturday
                saturday_routes = 0
                for route_name in applicable_routes:
                    if not route_name.upper().endswith('SA'):
                        continue
                    
                    if route_name not in route_definitions:
                        continue
                    
                    route_def = route_definitions[route_name]
                    vad_time = (route_def['vad_mit_schule'] if is_school_day
                               else route_def['vad_ohne_schule'])
                    
                    if not vad_time or vad_time == '00:00':
                        continue
                    
                    self.data['routes'].append({
                        'date': current_date,
                        'route_name': route_name,
                        'day_of_week': day_name,
                        'details': {
                            'type': 'saturday',
                            'duration_hours': route_def['diaten'] if route_def['diaten'] else 0,
                            'diaten': route_def['diaten'],
                            'vad_time': vad_time,
                            'location': route_def['kfz_ort'],
                            'season': season,
                            'school_status': school_status
                        }
                    })
                    saturday_routes += 1
                
                print(f"    ✅ Added {saturday_routes} SA routes")
            
            elif day_offset < 5:  # Weekdays
                weekday_routes = 0
                for route_name in applicable_routes:
                    if route_name.upper().endswith('SA'):
                        continue
                    
                    if route_name not in route_definitions:
                        if route_name in ['DI', 'MB']:
                            self.data['routes'].append({
                                'date': current_date,
                                'route_name': route_name,
                                'day_of_week': day_name,
                                'details': {
                                    'type': 'special_duty',
                                    'duty_code': route_name,
                                    'duty_name': self._get_duty_name(route_name),
                                    'season': season,
                                    'school_status': school_status
                                }
                            })
                            weekday_routes += 1
                        continue
                    
                    route_def = route_definitions[route_name]
                    vad_time = (route_def['vad_mit_schule'] if is_school_day
                               else route_def['vad_ohne_schule'])
                    
                    if not vad_time or vad_time == '00:00':
                        continue
                    
                    self.data['routes'].append({
                        'date': current_date,
                        'route_name': route_name,
                        'day_of_week': day_name,
                        'details': {
                            'type': 'regular',
                            'duration_hours': route_def['diaten'] if route_def['diaten'] else 0,
                            'diaten': route_def['diaten'],
                            'vad_time': vad_time,
                            'location': route_def['kfz_ort'],
                            'season': season,
                            'school_status': school_status
                        }
                    })
                    weekday_routes += 1
                
                print(f"    ✅ Added {weekday_routes} routes")
        
        print(f"✅ Total routes generated: {len(self.data['routes'])}")
    
    # ============= LENKER SHEET PARSING =============
    
    def parse_lenker_sheet(self):
        """Parse drivers sheet"""
        sheet = self._find_sheet('Lenker', 'Drivers', 'lenker', 'Feldkirchen')
        
        if not sheet:
            print("⚠️ Drivers sheet not found")
            return
        
        first_cell = sheet.cell(1, 1).value
        start_row = 2 if first_cell in ['Lenker', 'Name'] else 1
        
        for row_idx in range(start_row, sheet.max_row + 1):
            driver_name = sheet.cell(row_idx, 1).value
            
            if not driver_name or driver_name == '':
                break
            
            driver_name_str = str(driver_name).strip()
            if any(keyword in driver_name_str for keyword in [
                'Summe', 'Vollzeit', 'Feiertag', 'Krankenstand'
            ]):
                break
            
            col2 = sheet.cell(row_idx, 2).value
            col3 = sheet.cell(row_idx, 3).value
            col4 = sheet.cell(row_idx, 4).value
            col6 = sheet.cell(row_idx, 6).value
            col7 = sheet.cell(row_idx, 7).value
            col8 = sheet.cell(row_idx, 8).value
            
            soll_std = self._parse_time_to_hours(col2)
            b_grad = self._parse_percentage(col3)
            feiertag_hours = self._parse_time_to_hours(col4)
            krankenstand_hours = self._parse_time_to_hours(col6)
            fixdienst_ms = col7
            fixdienst_os = col8
            
            employment_type = self._determine_employment_type(b_grad)
            
            self.data['drivers'].append({
                'name': driver_name_str,
                'details': {
                    'type': employment_type,
                    'monthly_hours_target': soll_std,
                    'monthly_hours_worked': None,
                    'monthly_hours_remaining': None,
                    'feiertag_hours': feiertag_hours,
                    'krankenstand_hours': krankenstand_hours,
                    'employment_percentage': b_grad,
                    'fixed_route_with_school': str(fixdienst_ms).strip() if fixdienst_ms and fixdienst_ms != 'None' else None,
                    'fixed_route_without_school': str(fixdienst_os).strip() if fixdienst_os and fixdienst_os != 'None' else None
                }
            })
        
        print(f"✅ Parsed {len(self.data['drivers'])} drivers")
    
    # ============= FIXED ASSIGNMENTS PARSING =============
    
    def parse_fixed_assignments(self, week_start: date):
        """Parse fixed assignments day by day"""
        print("📌 Parsing fixed assignments (day by day)...")
        
        route_lookup = {}
        for route in self.data['routes']:
            key = (route['route_name'], route['date'])
            route_lookup[key] = route
        
        holiday_dates = set()
        for holiday in self.data['public_holidays']:
            if week_start <= holiday['date'] < week_start + timedelta(days=7):
                holiday_dates.add(holiday['date'])
        
        for driver_data in self.data['drivers']:
            driver_name = driver_data['name']
            fixdienst_ms = driver_data['details'].get('fixed_route_with_school')
            fixdienst_os = driver_data['details'].get('fixed_route_without_school')
            
            for day_offset in range(7):
                current_date = week_start + timedelta(days=day_offset)
                
                if current_date in holiday_dates:
                    continue
                
                is_school_day = self.data['school_days'].get(current_date, True)
                school_status = 'mit_schule' if is_school_day else 'ohne_schule'
                
                if is_school_day:
                    fixed_route_raw = fixdienst_ms
                    route_suffix = 'mS'
                else:
                    fixed_route_raw = fixdienst_os
                    route_suffix = 'oS'
                
                if not fixed_route_raw or fixed_route_raw == 'None' or str(fixed_route_raw).strip() == '':
                    continue
                
                fixed_route = str(fixed_route_raw).strip()
                
                if fixed_route.lower() == 'frei':
                    self.data['driver_availability'].append({
                        'driver_name': driver_name,
                        'date': current_date,
                        'available': False,
                        'notes': f'Fixdienst: frei ({school_status})'
                    })
                    continue
                
                if fixed_route in ['MB', 'DI']:
                    route_key = (fixed_route, current_date)
                    if route_key in route_lookup:
                        self.data['fixed_assignments'].append({
                            'driver_name': driver_name,
                            'route_name': fixed_route,
                            'date': current_date,
                            'notes': f'Special duty: {fixed_route}'
                        })
                    continue
                
                route_parts = [r.strip() for r in fixed_route.split('+')]
                primary_route_base = route_parts[0]
                primary_route_with_suffix = f"{primary_route_base}{route_suffix}"
                
                route_key = (primary_route_with_suffix, current_date)
                if route_key in route_lookup:
                    self.data['fixed_assignments'].append({
                        'driver_name': driver_name,
                        'route_name': primary_route_with_suffix,
                        'date': current_date,
                        'notes': f'Fixed assignment ({school_status})'
                    })
                else:
                    route_key_no_suffix = (primary_route_base, current_date)
                    if route_key_no_suffix in route_lookup:
                        self.data['fixed_assignments'].append({
                            'driver_name': driver_name,
                            'route_name': primary_route_base,
                            'date': current_date,
                            'notes': f'Fixed assignment ({school_status})'
                        })
        
        print(f"✅ Created {len(self.data['fixed_assignments'])} fixed assignments")
        print(f"✅ Created {len(self.data['driver_availability'])} 'frei' records")
    
    # ============= FEIERTAG SHEET PARSING =============
    
    def parse_feiertag_sheet(self):
        """Parse public holidays"""
        sheet = self._find_sheet('Feiertag', 'Holidays', 'feiertag', 'Feiertage')
        
        if not sheet:
            return
        
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            
            holiday_name = row[0].value
            holiday_date = row[1].value
            
            if isinstance(holiday_date, datetime):
                holiday_date = holiday_date.date()
            elif isinstance(holiday_date, str):
                holiday_date = self._parse_date(holiday_date)
            
            if holiday_date and holiday_name:
                self.data['public_holidays'].append({
                    'date': holiday_date,
                    'name': str(holiday_name).strip()
                })
        
        print(f"✅ Parsed {len(self.data['public_holidays'])} public holidays")
    
    # ============= HELPER METHODS =============
    
    def _determine_employment_type(self, percentage: Optional[int]) -> str:
        if not percentage:
            return "unknown"
        if percentage >= 100:
            return "full_time"
        elif percentage >= 80:
            return "reduced_hours"
        else:
            return "part_time"
    
    def _parse_time_to_hours(self, value) -> Optional[str]:
        if value is None:
            return None
        
        if isinstance(value, timedelta):
            total_seconds = int(value.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            return f"{hours:02d}:{minutes:02d}"
        
        if isinstance(value, datetime):
            return value.strftime('%H:%M')
        
        if isinstance(value, str):
            value = value.strip()
            if value == "00:00" or value == "":
                return "00:00"
            if ':' in value:
                return value
            return value
        
        if isinstance(value, (int, float)):
            hours = int(value)
            minutes = int((value - hours) * 60)
            return f"{hours:02d}:{minutes:02d}"
        
        return None
    
    def _subtract_time(self, time1: str, time2: str) -> str:
        try:
            h1, m1 = map(int, time1.split(':'))
            total_minutes1 = h1 * 60 + m1
            
            h2, m2 = map(int, time2.split(':'))
            total_minutes2 = h2 * 60 + m2
            
            result_minutes = total_minutes1 - total_minutes2
            
            if result_minutes < 0:
                result_minutes = 0
            
            hours = result_minutes // 60
            minutes = result_minutes % 60
            
            return f"{hours:02d}:{minutes:02d}"
        except:
            return "00:00"
    
    def _get_season_for_date(self, current_date: date) -> str:
        month = current_date.month
        if 6 <= month <= 9:
            return "summer"
        else:
            return "winter"
    
    def _get_season_key(self, season: str, school_status: str) -> str:
        return f"{season}_{school_status}"
    
    def _get_duty_name(self, code: str) -> str:
        mapping = {
            'MB': 'Mobilbüro',
            'DI': 'Dispo',
            'SOF': 'Sonderfahrt'
        }
        return mapping.get(code, code)
    
    def _parse_time(self, value) -> Optional[str]:
        if value is None:
            return None
        
        if isinstance(value, datetime):
            return value.strftime('%H:%M')
        
        if isinstance(value, str):
            if value == "00:00":
                return None
            return value.strip()
        
        return str(value)
    
    def _parse_number(self, value) -> Optional[float]:
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            try:
                return float(value.replace(',', '.'))
            except ValueError:
                return None
        
        return None
    
    def _parse_percentage(self, value) -> Optional[int]:
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return int(value)
        
        if isinstance(value, str):
            value = value.replace('%', '').strip()
            try:
                return int(float(value))
            except ValueError:
                return None
        
        return None
    
    def _parse_date(self, value) -> Optional[date]:
        if isinstance(value, datetime):
            return value.date()
        
        if isinstance(value, date):
            return value
        
        if isinstance(value, str):
            formats = [
                '%d-%m-%Y',
                '%d.%m.%Y',
                '%Y-%m-%d',
                '%d/%m/%Y'
            ]
            
            for fmt in formats:
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        
        return None